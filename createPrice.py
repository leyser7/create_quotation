import os
import sys
import random
import logging
import datetime
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from src.tkinter_app import tkinterApp
from src.cli_app import cliApp

# Initiate the parser
parser = argparse.ArgumentParser()

# Add long and short argument
parser.add_argument("--mode", "-m", help="set mode")

# Read arguments from the command line
args = parser.parse_args()

currentDateTime = datetime.datetime.now()
date = currentDateTime.date()
date = date.strftime("%d/%m/%Y")
logger = logging.getLogger()
logging.basicConfig(level=logging.INFO)


thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def set_border(ws, cell_range):
    rows = ws[cell_range]
    for row in rows:
        for r in row:
            r.border = thin_border


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


config_path = resource_path("config")
image_path = "tarjeta.png"
file_path = "cotizacion.xlsx"
output_file = "output.xlsx"
wb = load_workbook(f"{config_path}/{file_path}")

ws = wb["Hoja1"]  # or wb.active

ws["R4"] = f"A-{random.randint(10150,10300)}"
ws.merge_cells("R4:U4")
set_border(ws, "R4:U4")
ws["R4"].alignment = Alignment(horizontal="center")


# Check for mode
if args.mode == "cli":
    app = cliApp()
    app.exec()
else:
    app = tkinterApp()
    app.mainloop()


client_name = app.client_name
description = app.description
items_n = app.items_n
unit_values = app.unit_values
items_desc = app.items_desc

####
ws["M11"] = date
ws["D11"] = client_name
ws["A22"] = description
for n, (item_n, unit_value, item_desc) in enumerate(
    zip(items_n, unit_values, items_desc)
):
    try:
        ws[f"L{14 + n}"] = item_n
        ws[f"O{14 + n}"] = unit_value
        ws[f"C{14 + n}"] = item_desc
        ws[f"C{14 + n}"].alignment = Alignment(
            horizontal="center", wrap_text=True
        )
        height = len(ws[f"C{13 + n}"].value) / 3
        ws.row_dimensions[14 + n].height = 13.8 if height < 13.8 else height
    except ValueError:
        logger.error("[Error] No se puede registrando el item")

####
wb.save(output_file)
os.system(
    "libreoffice --headless --convert-to pdf " f"{output_file} >/dev/null 2>&1"
)
logger.info(f"{output_file} generated")
